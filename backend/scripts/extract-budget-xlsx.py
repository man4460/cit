# -*- coding: utf-8 -*-
"""Extract budget rows from ฝรภ. Excel → JSON stdout"""
import json
import sys

import openpyxl

path = sys.argv[1]
wb = openpyxl.load_workbook(path, data_only=True)


def cell(row, idx):
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    return s if s else None


def num(v):
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def base(i, kind):
    return {
        "row": i,
        "kind": kind,
        "note": None,
        "codeOrSection": None,
        "fileRef": None,
        "nameA": None,
        "nameB": None,
        "definition": None,
        "superiorCi": None,
        "buyerName": None,
        "requestingUnit": None,
        "y69": None,
        "carryIn": None,
        "spent": None,
        "y70": None,
        "delta": None,
        "deltaPct": None,
        "dir": None,
        "reason": None,
        "y71": None,
        "commitReq70": None,
        "commitAppr70": None,
    }


request_rows = []
ws = wb["คำขอปึ70"]
kind = "EXPENSE"
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i < 5:
        continue
    code = cell(row, 3)
    if code and "สินทรัพย์ถาวร" in str(code):
        kind = "CAPEX"
    if code and "หมวดค่าใช้จ่าย" in str(code):
        kind = "EXPENSE"
    name_a, name_b, definition, file_ref = cell(row, 5), cell(row, 6), cell(row, 7), cell(row, 4)
    if not any([code, name_a, name_b, definition, file_ref, cell(row, 11)]):
        continue
    r = base(i, kind)
    r.update(
        {
            "note": None if cell(row, 1) is None else str(cell(row, 1)),
            "codeOrSection": None if code in (None, "--") else str(code),
            "fileRef": None if file_ref is None else str(file_ref),
            "nameA": None if name_a is None else str(name_a),
            "nameB": None if name_b is None else str(name_b),
            "definition": None if definition is None else str(definition),
            "y69": num(cell(row, 11)),
            "spent": num(cell(row, 12)),
            "y70": num(cell(row, 14)),
            "delta": num(cell(row, 15)),
            "deltaPct": num(cell(row, 16)),
            "dir": None if cell(row, 17) is None else str(cell(row, 17)),
            "reason": None if cell(row, 18) is None else str(cell(row, 18)),
            "y71": num(cell(row, 20)),
        }
    )
    request_rows.append(r)

spend_rows = []
ws2 = wb["สรุปใช้งบ 31กค"]
kind2 = "EXPENSE"
for i, row in enumerate(ws2.iter_rows(values_only=True), 1):
    if i < 10:
        continue
    superior, code, name, definition = cell(row, 0), cell(row, 1), cell(row, 2), cell(row, 3)
    if superior and "สินทรัพย์ถาวร" in str(superior):
        kind2 = "CAPEX"
    if not any([superior, code, name, definition]):
        continue
    if str(superior or "").startswith("รวม"):
        continue
    r = base(i, kind2)
    r.update(
        {
            "codeOrSection": None if code is None else str(code),
            "nameA": None if name is None else str(name),
            "definition": None if definition is None else str(definition),
            "superiorCi": None if superior is None else str(superior),
            "buyerName": None if cell(row, 5) is None else str(cell(row, 5)),
            "requestingUnit": None if cell(row, 4) is None else str(cell(row, 4)),
            "y69": num(cell(row, 7)),
            "carryIn": num(cell(row, 6)),
            "spent": num(cell(row, 11)),
        }
    )
    spend_rows.append(r)

result_rows = []
ws3 = wb["ผลพิ69-สรุป"]
kind3 = "EXPENSE"
for i, row in enumerate(ws3.iter_rows(values_only=True), 1):
    if i < 11:
        continue
    superior, code, name, definition = cell(row, 0), cell(row, 1), cell(row, 2), cell(row, 3)
    if superior and ("สินทรัพย์ถาวร" in str(superior) or "หมวด 80" in str(superior)):
        kind3 = "CAPEX"
    if not any([code, name]) and not (superior and str(superior).isdigit()):
        continue
    if str(superior or "").startswith("รวม"):
        continue
    r = base(i, kind3)
    r.update(
        {
            "superiorCi": None if superior is None else str(superior),
            "codeOrSection": None if code is None else str(code),
            "nameA": None if name is None else str(name),
            "definition": None if definition is None else str(definition),
            "y69": num(cell(row, 8)) if num(cell(row, 8)) is not None else num(cell(row, 5)),
            "commitReq70": num(cell(row, 6)),
            "commitAppr70": num(cell(row, 10)),
        }
    )
    result_rows.append(r)

print(json.dumps({"requestRows": request_rows, "spendRows": spend_rows, "resultRows": result_rows}, ensure_ascii=False))
